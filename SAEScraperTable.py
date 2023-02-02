import openpyxl as xl
import re

global ArmGroupLabels
global Type
global OtherName
global Name
global Description
ArmGroupLabels = 13
Type = 14
OtherName = 15
Name = 16
Description = 17


def estIdentifiant(string):
    rep = False
    if string == "arm_group_labels":
        rep = True
    elif string == "type":
        rep = True
    elif string == "other_names":
        rep = True
    elif string == "name":
        rep = True
    elif string == "description":
        rep = True
    return rep

def getNumString(string):
    rep = 18
    if string == "arm_group_labels":
        rep = ArmGroupLabels
    elif string == "type":
        rep = Type
    elif string == "other_names":
        rep = OtherName
    elif string == "name":
        rep = Name
    elif string == "description":
        rep = Description
    return rep

path = "D:\\IUT\\Semestre4\\SAE R4.C10\\20200601_IRIT_clinicalTrials+publications.xlsx"

file = xl.load_workbook(path)

sheet1 = file["1 - ClinicalTrials_ObsStudies"]
sheet2 = file["2 - ClinicalTrials_RandTrials"]

##cell_obj = sheet1.cell(row = 2, column = 1)
##sheet1.cell(row=2, column=18).value="TEST"
##rep = re.findall(r"'(.*?)'", sheet1.cell(row = 2, column = 8).value)
##print(rep)

for i in range(2, sheet1.max_row):
    string = sheet1.cell(row = i, column = 8).value
    if not string == None:
        decoupe = re.findall(r"'(.*?)'", string)
        for j in range (len(decoupe)):
            if estIdentifiant(decoupe[j]):
                if j+1 < len(decoupe):
                    if not estIdentifiant(decoupe[j+1]):
                        data = decoupe[j+1]
                        num = getNumString(decoupe[j])
                        sheet1.cell(row=i, column=num).value=data

for i in range(2, sheet2.max_row):
    string = sheet2.cell(row = i, column = 8).value
    if not string == None:
        decoupe = re.findall(r"'(.*?)'", string)
        for j in range (len(decoupe)):
            if estIdentifiant(decoupe[j]):
                if j+1 < len(decoupe):
                    if not estIdentifiant(decoupe[j+1]):
                        data = decoupe[j+1]
                        num = getNumString(decoupe[j])
                        sheet2.cell(row=i, column=num).value=data

file.save(path)