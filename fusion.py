import openpyxl as xl

def IdInSheet(id, sheetOut, num) :
    for m in range(1, sheetOut.max_row):
        if id == sheetOut.cell(m, 1).value :
            sheetOut.cell(m, num).value = True 
            return True
    return False

pathin = "D:\\IUT\\Semestre4\\SAE R4.C10\\20200601_IRIT_clinicalTrials+publications.xlsx"
pathout = "D:\\IUT\\Semestre4\\SAE R4.C10\\TableSAEFusion1.5.xlsx"

filein = xl.load_workbook(pathin)
fileout = xl.load_workbook(pathout)

sheetInTrialObs = filein["1 - ClinicalTrials_ObsStudies"]
sheetInTrialRand = filein["2 - ClinicalTrials_RandTrials"]
sheetInPubliObs = filein["3 - Publications_ObsStudies"]
sheetInPubliRand = filein["4 - Publications_RandTrials"]

sheetOutTrial = fileout["ClinicalTrials"]
sheetOutPubli = fileout["Publications"]

counter = 2

for i in range(2, sheetInTrialObs.max_row):
    if sheetInTrialObs.cell(i,1).value != None:
        for j in range(1, sheetInTrialObs.max_column):
            if j < 8:
                buffer = sheetInTrialObs.cell(i,j).value
                sheetOutTrial.cell(counter, j).value = buffer
            if j > 8:
                buffer = sheetInTrialObs.cell(i,j).value
                sheetOutTrial.cell(counter, j-1).value = buffer
        sheetOutTrial.cell(counter, 17).value = True
        sheetOutTrial.cell(counter, 18).value = False
        counter = counter + 1


for i in range(2, sheetInTrialRand.max_row):
    if sheetInTrialRand.cell(i,1).value != None:
        rep = IdInSheet(sheetInTrialRand.cell(i,1).value, sheetOutTrial, 18)
        if not rep :
            for j in range(1, sheetInTrialRand.max_column):
                buffer = sheetInTrialRand.cell(i,j).value
                sheetOutTrial.cell(counter, j).value = buffer
            sheetOutTrial.cell(counter, 18).value = True
            sheetOutTrial.cell(counter, 17).value = False
            counter = counter + 1
    
counter = 2

for i in range(2, sheetInPubliObs.max_row):
    if sheetInPubliObs.cell(i,1).value != None:
        for j in range(1, 18):
            buffer = sheetInPubliObs.cell(i,j).value
            sheetOutPubli.cell(counter, j).value = buffer
        sheetOutPubli.cell(counter, 18).value = True
        sheetOutPubli.cell(counter, 19).value = False
        counter = counter + 1

for i in range(2, sheetInPubliRand.max_row):
    if sheetInPubliRand.cell(i,1).value != None:
        rep = IdInSheet(sheetInPubliRand.cell(i,1).value, sheetOutPubli, 19)
        if not rep :
            for j in range(1, 18):
                buffer = sheetInPubliRand.cell(i,j).value
                sheetOutPubli.cell(counter, j).value = buffer
            sheetOutPubli.cell(counter, 19).value = True
            sheetOutPubli.cell(counter, 18).value = False
            counter = counter + 1

fileout.save(pathout)