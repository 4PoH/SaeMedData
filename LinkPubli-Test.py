import openpyxl as xl

urlPubli = "D:\\IUT\\Semestre4\\SAE R4.C10\\TableSAEFusion2.0.xlsx"

file = xl.open(urlPubli)
sheetPub = file["Publications"]
sheetLink = file["Link"]

buffer = 2

for i in range(2,sheetPub.max_row):
    valeur = sheetPub.cell(i,17).value
    if (valeur != None):
        liste = valeur.split("/")
        for id in liste:
            sheetLink.cell(buffer, 1).value = sheetPub.cell(i,1).value
            sheetLink.cell(buffer, 2).value = id
            buffer = buffer + 1

print (buffer)
file.save(urlPubli)