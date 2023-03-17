import requests, PyPDF2
from io import BytesIO
from selenium import webdriver
import openpyxl as xl

def getStringFromUrl(url):
    response = requests.get(url)
    rawData = response.content
    with BytesIO(rawData) as data:
        try :
            readPdf = PyPDF2.PdfReader(data)
            for page in range(len(readPdf.pages)):
                string = readPdf.pages[page].extract_text().encode('utf8')
        except:
            driver = webdriver.Chrome()
            driver.get(url)
            string = driver.page_source.encode('utf8')
            driver.quit()
    string = string.decode('utf8')
    return(string)

def parcourTest(sheetPubli, sheetTest, ligne, string) :
    boolRep = False
    for y in range(2, sheetTest.max_row):
        nom = sheetTest.cell(row=y, column=10).value
        id = sheetTest.cell(row=y, column=1).value
        if id is not None :
            valeur = sheetPubli.cell(row=ligne, column=17).value
            if id in string :
                if valeur is None:
                    sheetPubli.cell(row=ligne, column=17).value = id
                else :
                    if id not in valeur :
                        sheetPubli.cell(row=ligne, column=17).value = sheetPubli.cell(row=ligne, column=17).value + ' / ' + id
                        boolRep = True
            elif nom is not None:
                if nom in string:
                    if valeur is None:
                        sheetPubli.cell(row=ligne, column=17).value = id
                    else :
                        if id not in valeur :
                            sheetPubli.cell(row=ligne, column=17).value = sheetPubli.cell(row=ligne, column=17).value + ' / ' + id
                            boolRep = True
    return boolRep


path = "D:\\IUT\\Semestre4\\SAE R4.C10\\20200601_IRIT_clinicalTrials+publications.xlsx"

file = xl.load_workbook(path)

sheet1 = file["1 - ClinicalTrials_ObsStudies"]
sheet2 = file["2 - ClinicalTrials_RandTrials"]
sheet3 = file["3 - Publications_ObsStudies"]
sheet4 = file["4 - Publications_RandTrials"]
counter = 0

# #Parcours Publi Obsv
for i in range(2, sheet3.max_row):
    url = sheet3.cell(i ,8).value
    if url is not None:
        print(i , " " , url)
        string = getStringFromUrl(url)
        if parcourTest(sheetPubli=sheet3 ,sheetTest=sheet1 ,ligne=i ,string=string):
            counter = counter + 1
            print(counter)
            file.save(path)
        if parcourTest(sheetPubli=sheet3 ,sheetTest=sheet2 ,ligne=i ,string=string):
            counter = counter + 1
            print(counter)
            file.save(path)

#Parcours Publi Rand
for i in range(2, sheet4.max_row):
    url = sheet4.cell(i ,8).value
    if url is not None:
        print(i, " ", url)
        if url != "https://pubs.rsc.org/en/content/articlepdf/2020/ra/d0ra03582c":
            string = getStringFromUrl(url)
            if parcourTest(sheetPubli=sheet4 ,sheetTest=sheet1 ,ligne=i, string=string):
                counter = counter + 1
                print(counter)
                file.save(path)
            if parcourTest(sheetPubli=sheet4 ,sheetTest=sheet2 ,ligne=i ,string=string):
                counter = counter + 1
                print(counter)
                file.save(path)
print("On à trouver un total de " , counter , " liaison(s)")
file.save(path)
