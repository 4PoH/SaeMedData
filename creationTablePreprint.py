import openpyxl as xl
import csv

urlPubli = "D:\\IUT\\Semestre4\\SAE R4.C10\\TableSAEFusion2.0.xlsx"
urldata = "D:\IUT\Semestre4\SAE R4.C10\preprint_tracker.csv"

file = xl.load_workbook(urlPubli)
sheetPub = file["Publications"]
sheetTest = file["ClinicalTrials"]
sheetPreprint = file["Preprint"]

i= 2

with open(urldata) as fileData:
    reader = csv.DictReader(fileData)
    for row in reader:
        sheetPreprint.cell(row=i, column=1).value = row["Preprint Server"]
        sheetPreprint.cell(row=i, column=2).value = row["Preprint DOI"]
        sheetPreprint.cell(row=i, column=3).value = int(row["Latest Version"])
        sheetPreprint.cell(row=i, column=4).value = row["Date of Latest Version"]
        sheetPreprint.cell(row=i, column=5).value = row["Title"]
        sheetPreprint.cell(row=i, column=6).value = row["Byline"]
        if (row["Publication DOI"]=="-"):
            sheetPreprint.cell(row=i, column=7).value = None
        else:
            sheetPreprint.cell(row=i, column=7).value = row["Publication DOI"]
        sheetPreprint.cell(row=i, column=8).value = row["Status"]
        if (row["Assessors"]=="-"):
            sheetPreprint.cell(row=i, column=9).value = None
        else:
            sheetPreprint.cell(row=i, column=9).value = row["Assessors"]
        i=i+1

file.save(urlPubli)